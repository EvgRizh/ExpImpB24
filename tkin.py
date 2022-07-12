import tkinter
from tkinter import *
from tkinter import filedialog, ttk
from openpyxl import Workbook, load_workbook
from fast_bitrix24 import Bitrix


def import_to_bitrix():
    wbi = load_workbook(entry.get())
    wsi = wbi.active
    xl_tuple = []
    for row in wsi.iter_rows(min_row=2, values_only=True):
        xl_tuple.append(row)
    print(xl_tuple)
    params = []
    param = {}
    for tup in xl_tuple:
        if tup[0] == None:
            n = 0
            param.clear()
            fields = {}
            for k in fields_to_xls.keys():
                if k == "MEASURE":
                    m = tup[n]
                    for d in measure:
                        if m == d["SYMBOL_RUS"]:
                            fields.update({"MEASURE": d["ID"]})
                            n = n+1
                            break
                        elif m == None or m == "отсутствует":
                            fields.update({"MEASURE": None})
                            n = n+1
                            break
                elif k == "CATALOG_1":
                    if tup[n] in dict_section_name.keys():
                        if tup[n+1] == None:
                            fields.update({"SECTION_ID": dict_section_name[tup[n]]})
                            n = n + 1
                            break
                        else:
                            if tup[n+1] in dict_section_name.keys() and dict_name_sectionid[tup[n+1]] == dict_section_name[tup[n]]:
                                if tup[n+2] == None:
                                    fields.update({"SECTION_ID": dict_section_name[tup[n+1]]})
                                    break
                                else:
                                    if tup[n+2] == y["NAME"] and y["SECTION_ID"] == x["ID"]:
                                        fields.update({"SECTION_ID": y["ID"]})
                                        break
                    # elif tup[n] not in dict_section_name:
                    #     print(tup[n])
                    #     b.call("crm.productsection.add", {"CATALOG_ID": "25", "NAME": tup[n], "SECTION_ID": None})
                    #     list_section = b.get_all("crm.productsection.list")
                    #     for list in list_section:
                    #         if list["NAME"] == tup[n]:
                    #             section_id = list["ID"]
                    #             fields.update({"SECTION_ID": section_id})
                else:
                    fields.update({k: tup[n]})
                    n = n+1
            param.update({"fields": fields})
        else:
            pass
        params.append(param)

    # b.call("crm.product.add", params)
    print(params)
    button_import["state"] = tkinter.DISABLED

def export_from_bitrix():
    pos = b.get_all("crm.product.list")
    print(pos)
    dict_post = {}
    for dict in pos:
        for k in list(fields_to_xls.keys()):
            if k == "MEASURE":
                for d in measure:
                    if dict[k] == None:
                        dict_post[k] = "отсутствует"
                    if d["ID"] == dict[k]:
                        dict_post[k] = d["SYMBOL_RUS"]
            elif k == "CATALOG_1":
                if dict["SECTION_ID"] != None:
                    if dict_section_link[dict["SECTION_ID"]] == None:
                        dict_post["CATALOG_1"] = dict_section[dict["SECTION_ID"]]
                        dict_post["CATALOG_2"] = None
                        dict_post["CATALOG_3"] = None
                        continue
                    elif dict_section_link[dict_section_link[dict["SECTION_ID"]]]==None:
                        dict_post["CATALOG_2"] = dict_section[dict["SECTION_ID"]]
                        dict_post["CATALOG_1"] = dict_section[dict_section_link[dict["SECTION_ID"]]]
                        dict_post["CATALOG_3"] = None
                        continue
                    elif dict_section_link[dict_section_link[dict_section_link[dict["SECTION_ID"]]]] == None:
                        dict_post["CATALOG_1"] = dict_section[dict_section_link[dict_section_link[dict["SECTION_ID"]]]]
                        dict_post["CATALOG_2"] = dict_section[dict_section_link[dict["SECTION_ID"]]]
                        dict_post["CATALOG_3"] = dict_section[dict["SECTION_ID"]]
                        continue
                else:
                    dict_post["CATALOG_1"] = None
                    dict_post["CATALOG_2"] = None
                    dict_post["CATALOG_3"] = None
                    continue
            elif k in ["CATALOG_2", "CATALOG_3"]:
                continue
            else:
                dict_post[k] = dict[k]

        ws.append(list(dict_post.values()))
    wb.save(entry_save.get() + ".xlsx")
    label_export.config(text="Товары экспортированы\nв файл {}.xlsx".format(entry_save.get()))
    button_export["state"] = tkinter.DISABLED



def openfile():
    entry.delete(0, END)
    path = filedialog.askopenfilename()
    entry.insert(0, path)
    if len(entry.get()) > 0:
        button_import["state"] = tkinter.ACTIVE


def set_file_save():
    button_export["state"] = tkinter.ACTIVE

webhook = "https://b24-b7j74c.bitrix24.ru/rest/1/jznpvcpgth3dlisy/"
b = Bitrix(webhook)

file_name = ""
section = b.get_all("crm.productsection.list")
print(section)
measure = b.get_all("crm.measure.list")
dict_section = {}
dict_section_link = {}
dict_section_name = {}
dict_name_sectionid = {}
for dict in section:
    dict_section_name[dict["NAME"]] = dict["ID"]
    dict_name_sectionid[dict["NAME"]] = dict["SECTION_ID"]
    dict_section[dict["ID"]] = dict["NAME"]
    dict_section_link[dict["ID"]] = dict["SECTION_ID"]
interface = Tk()
interface.geometry("640x480")
interface.title("Экспорт/импорт товаров Битрикс24")
interface.columnconfigure(0, weight=1)
interface.rowconfigure(0, weight=1)
interface.rowconfigure(1, weight=1)
frame1 = LabelFrame(interface, bg="grey", text="Импорт товаров в Битрикс24")
frame2 = LabelFrame(interface, bg="grey", text="Экспорт товаров из Битрикс24")
frame1.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
frame2.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
button = ttk.Button(frame1, text="Open", command=openfile)
button_import = ttk.Button(frame1, text="Import", command=import_to_bitrix, state=tkinter.DISABLED)
button_export = ttk.Button(frame2, text="Export", command=export_from_bitrix, state=tkinter.DISABLED)
button_save = ttk.Button(frame2, text="Назначить имя", command=set_file_save)
label_file_save = ttk.Label(frame2, background="gray", text="Имя файла для сохранения\n(пишите без расширения):")
label_export = ttk.Label(frame2, background="grey")
entry = ttk.Entry(frame1, width=60)
entry_save = ttk.Entry(frame2, width=20)
entry.grid(column=1, row=0, padx=10, pady=10, sticky="w")
button.grid(column=0, row=0, padx=10, pady=10, sticky="w")
button_import.grid(column=0, row=1, padx=10, pady=10, sticky="w")
label_file_save.grid(column=0, row=0, padx=10, pady=10, sticky="w")
entry_save.grid(column=0, row=1, padx=10, pady=10, sticky="w")
button_save.grid(column=1, row=1, padx=10, pady=10, sticky="w")
button_export.grid(column=0, row=2, padx=10, pady=10, sticky="w")
label_export.grid(column=0, row=3, padx=10, pady=10, sticky="w")


fields_to_xls = {"ID": "ID", "XML_ID": "Внешний код", "NAME": "Название",
                 "CODE": "Символьный код", "DESCRIPTION": "Описание",
                 "DESCRIPTION_TYPE": "Формат описания", "ACTIVE": "Активен",
                 "CURRENCY_ID": "Валюта", "PRICE": "Цена", "MEASURE": "Единица измерения",
                 "CATALOG_1": "Каталог(уровень 1)", "CATALOG_2": "Каталог(уровень 2)", "CATALOG_3": "Каталог(уровень 3)"}

wb = Workbook()
ws = wb.active
ws.append(list(fields_to_xls.values()))

interface.mainloop()